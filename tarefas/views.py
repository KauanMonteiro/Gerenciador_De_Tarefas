from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404 , redirect
from django.urls import reverse 
from .models import Tarefas, Equipe
from usuario.models import Usuario
import matplotlib.pyplot as plt
from io import BytesIO
import base64
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def home(request):
    if 'usuario' not in request.session:
        return redirect('login')

    usuario_id = request.session['usuario']
    usuario = Usuario.objects.get(pk=usuario_id)
    
    if usuario.equipes.exists():
        equipe_usuario = usuario.equipes.first()
        tarefas_incompletas = Tarefas.objects.exclude(concluida=usuario)
        return render(request, 'tarefas/pages/home.html', {'tarefas': tarefas_incompletas, 'usuario': usuario})
    else:
        return render(request, 'tarefas/pages/home.html')


def tarefa_detail(request, id):
    if 'usuario' not in request.session:
        return redirect('login')
    tarefa = get_object_or_404(Tarefas, pk=id)
    return render(request, 'tarefas/pages/tarefa_detalhe.html', {'tarefa': tarefa, 'is_detail_page': True})

def concluir_tarefa(request, id):
    if 'usuario' not in request.session:
        return redirect('login')

    usuario_id = request.session['usuario']
    usuario = Usuario.objects.get(pk=usuario_id)

    tarefa = get_object_or_404(Tarefas, pk=id)

    if usuario in tarefa.tarefa_para.first().membros.all():
        if not tarefa.concluida.filter(pk=usuario_id).exists():
            tarefa.concluida.add(usuario)

    if request.method == 'POST':
        resposta_usuario = request.POST.get('alternativa')
        tarefa.resposta_usuario = resposta_usuario
        tarefa.save()
        
    return redirect(reverse('home'))

def area_usuario(request):
    if 'usuario' not in request.session:
        return redirect('login')
    
    usuario_id = request.session['usuario']
    usuario = Usuario.objects.get(pk=usuario_id)
    estatisticas_usuario = []
    if usuario.equipes.exists():
        equipes_usuario = usuario.equipes.all()
      
        
        for equipe in equipes_usuario:
            tarefas_equipe = Tarefas.objects.filter(tarefa_para=equipe)
            
            tarefas_nao_concluidas = tarefas_equipe.exclude(concluida=usuario).count()
            tarefas_concluidas = tarefas_equipe.filter(concluida=usuario).count()
            total_tarefas = tarefas_equipe.count()
            
            estatisticas_usuario.append({
                'equipe': equipe,
                'total_tarefas': total_tarefas,
                'tarefas_concluidas': tarefas_concluidas,
                'tarefas_nao_concluidas': tarefas_nao_concluidas,
                })
            
        return render(request, 'tarefas/pages/area_usuario.html', {
            'usuario': usuario,
            'equipes': equipes_usuario,
            'estatisticas_usuario': estatisticas_usuario,
        })
    else:
        return render(request, 'tarefas/pages/home.html')


def criar_equipe(request):
    if 'usuario' not in request.session:
        return redirect('login')
    
    usuario_id = request.session['usuario']
    usuario = Usuario.objects.get(pk=usuario_id)
    
    if usuario.cargo != 'Professor':
        return render(request, 'tarefas/pages/home.html')
    else:
        if request.method == 'POST':
            nome = request.POST.get('nome')

            if len(nome.strip()) == 0:
                return render(request, 'tarefas/pages/criar_equipe.html')
            
            else:
                equipe = Equipe(nome_equipe=nome)
                equipe.save()
                return redirect(reverse('home'))
        
        return render(request, 'tarefas/pages/criar_equipe.html')

def equipe_detalhe(request, id):
    if 'usuario' not in request.session:
        return redirect('login')
    
    usuario_id = request.session['usuario']
    usuario = Usuario.objects.get(pk=usuario_id)
    
    equipe = get_object_or_404(Equipe, pk=id)
    tarefas_equipe = Tarefas.objects.filter(tarefa_para=equipe)
    
    return render(request, 'tarefas/pages/ver_mais_equipe.html', {'equipe': equipe, 'tarefas': tarefas_equipe, 'usuario': usuario})
    

def adicao_usuario(request, equipe_id):
    if 'usuario' not in request.session:
        return redirect('login')
    
    if request.method == 'POST':
        userId = request.POST.get('userId')

        if len(userId.strip()) == 0:
            return redirect('adicao_usuario', equipe_id=equipe_id)
        
        else:
            try:
                equipe = Equipe.objects.get(id=equipe_id)
                usuario = Usuario.objects.get(id=userId)
                equipe.membros.add(usuario)
                equipe.save()
                return redirect('area_usuario')
            except Equipe.DoesNotExist:
                pass

    return render(request, 'tarefas/pages/adicionar_usuario.html',{'equipe_id': equipe_id})

def adicionar_tarefas(request):
    if 'usuario' not in request.session:
        return redirect('login')
    
    usuario_id = request.session['usuario']
    usuario = Usuario.objects.get(pk=usuario_id)
    equipe = usuario.equipes.all()
    
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        descricao = request.POST.get('descricao')
        data_limite = request.POST.get('data_limite')
        tarefa_para = request.POST.getlist('tarefa_para')
        alternativa1 = request.POST.get('alternativa1')
        alternativa2 = request.POST.get('alternativa2')
        alternativa3 = request.POST.get('alternativa3')
        alternativa4 = request.POST.get('alternativa4')
        alternativa5 = request.POST.get('alternativa5')
        alternativa_correta = request.POST.get('alternativa_correta')

        if not (titulo and descricao and tarefa_para):
            return redirect('adicionar_tarefas')
        else:
            tarefa = Tarefas.objects.create(titulo=titulo,
                    descricao=descricao,
                    data_limite=data_limite,
                    autor=usuario,alternativa1=alternativa1,
                    alternativa2 =alternativa2,
                    alternativa3 = alternativa3,
                    alternativa4 = alternativa4,
                    alternativa5 = alternativa5 ,
            alternativa_correta=alternativa_correta)
            
            tarefa.tarefa_para.set(tarefa_para)
            tarefa.save()
            return redirect('home')
    return render(request, 'tarefas/pages/adicionar_tarefa.html', {'equipes': equipe})

def estatisticas(request, equipe_id):
    equipe = get_object_or_404(Equipe, pk=equipe_id)
    membros_equipe = equipe.membros.all()
    estatisticas_usuarios = []
    estatisticas_equipe = {
        'tarefas_concluidas': 0,
        'tarefas_nao_concluidas': 0,
        'corretas': 0,
        'erradas': 0,
    }
    for usuario in membros_equipe:
        tarefas_usuario = Tarefas.objects.filter(tarefa_para=equipe)
        tarefas_concluidas = tarefas_usuario.filter(concluida=usuario)
        tarefas_nao_concluidas = tarefas_usuario.exclude(concluida=usuario).count()
 
        corretas = 0
        erradas = 0
        for tarefa in tarefas_concluidas:
            if tarefa.resposta_usuario != tarefa.alternativa_correta:
                erradas += 1
            else:
                corretas += 1

        estatisticas_equipe['tarefas_concluidas'] += tarefas_concluidas.count()
        estatisticas_equipe['tarefas_nao_concluidas'] += tarefas_nao_concluidas

        labels = ['Concluídas', 'Não Concluídas']
        sizes = [estatisticas_equipe['tarefas_concluidas'], estatisticas_equipe['tarefas_nao_concluidas']]


        plt.figure(figsize=(7.5, 7))
        plt.pie(sizes, labels=labels, autopct='%1.1f%%', colors=['green','red'],startangle=140)
        plt.axis('equal')

        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()

        grafico = base64.b64encode(image_png).decode('utf-8')

        labels = ['Concluídas', 'Não Concluídas']
        sizes = [tarefas_concluidas.count(), tarefas_nao_concluidas]

        plt.figure(figsize=(7.5, 7))
        plt.pie(sizes, labels=labels, autopct='%1.1f%%', colors=['green','red'], startangle=140)
        plt.axis('equal')

        buffer = BytesIO()
        plt.savefig(buffer, format='jpg')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()

        grafico_usuario = base64.b64encode(image_png).decode('utf-8')

        estatisticas_usuarios.append({
            'usuario': usuario,
            'tarefas_concluidas': tarefas_concluidas.count(),
            'tarefas_nao_concluidas': tarefas_nao_concluidas,
            'corretas': corretas,
            'erradas': erradas,
            'grafico_usuario': grafico_usuario,  
        })
    

    return render(request, 'tarefas/pages/estatisticas.html', {'equipe': equipe, 'estatisticas_usuarios': estatisticas_usuarios,'equipe': equipe,
        'estatisticas_equipe': estatisticas_equipe,
        'grafico': grafico,})


def download_excel(request, equipe_id):
    equipe = get_object_or_404(Equipe, pk=equipe_id)
    membros_equipe = equipe.membros.all()
    estatisticas_usuarios = []
    wb = Workbook()
    ws = wb.active
    ws.title = 'Estatisticas Equipe'
    
    ws['A1'] = 'Usuário'
    ws['B1'] = 'Tarefas Concluídas'
    ws['C1'] = 'Tarefas Não Concluídas'
    ws['D1'] = 'Corretas'
    ws['E1'] = 'Erradas'

    row_num = 2
    for usuario in membros_equipe:
        tarefas_usuario = Tarefas.objects.filter(tarefa_para=equipe)
        tarefas_concluidas = tarefas_usuario.filter(concluida=usuario)
        tarefas_nao_concluidas = tarefas_usuario.exclude(concluida=usuario)


        corretas = 0
        erradas = 0
        for tarefa in tarefas_concluidas:
            if tarefa.resposta_usuario != tarefa.alternativa_correta:
                erradas += 1
            else:
                corretas += 1
        ws[f'A{row_num}'] = usuario.nome 
        ws[f'B{row_num}'] = tarefas_concluidas.count()
        ws[f'C{row_num}'] = tarefas_nao_concluidas.count()
        ws[f'D{row_num}'] = corretas
        ws[f'E{row_num}'] = erradas
        
        row_num += 1

    arquivo_excel = BytesIO()
    wb.save(arquivo_excel)
    arquivo_excel.seek(0)

    response = HttpResponse(arquivo_excel, content_type='application/vnd.openpyxl.sheet')
    response['Content-Disposition'] = 'attachment; filename=EstatisticasEquipe.xlsx'

    return response